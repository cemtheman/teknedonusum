"""XLSX export of existing decision-summary and transparency rows."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from calculations.assumptions_transparency import AssumptionSourceRow
from calculations.decision_summary import VesselDecisionSummaryRow
from models.compliance import ComplianceStatus


DECISION_HEADERS = (
    "Tekne tipi",
    "Yolcu kapasitesi",
    "Seçilen hız (knot)",
    "Batarya kapasitesi (kWh)",
    "Günlük sevk enerjisi (kWh/gün)",
    "Güneş katkısı (kWh/gün)",
    "Net şebeke ihtiyacı (kWh/gün)",
    "Tahmini menzil (NM)",
    "Teknik uygunluk",
    "Yatırım maliyeti (TL)",
    "Hibe (TL)",
    "Net yatırım (TL)",
    "Yıllık işletme tasarrufu (TL)",
    "Geri ödeme süresi (sezon)",
    "Yıllık CO₂ azaltımı (ton/yıl)",
)

ASSUMPTION_HEADERS = (
    "Parametre",
    "Mevcut değer",
    "Kaynak türü",
    "Açıklama",
)

STATUS_LABELS = {
    ComplianceStatus.PASS: "Uygun",
    ComplianceStatus.FAIL: "Uygun değil",
    None: "Henüz değerlendirilmedi",
}


def _style_sheet(sheet, widths):
  header_fill = PatternFill("solid", fgColor="1E3A8A")
  header_font = Font(bold=True, color="FFFFFF")
  for cell in sheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  sheet.freeze_panes = "A2"
  sheet.auto_filter.ref = sheet.dimensions
  sheet.row_dimensions[1].height = 32
  for column_letter, width in widths.items():
    sheet.column_dimensions[column_letter].width = width
  for row in sheet.iter_rows(min_row=2):
    for cell in row:
      cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_decision_summary_xlsx(decision_rows, assumption_rows) -> bytes:
  """Export current structured rows without recalculating their values."""
  decision_rows = tuple(decision_rows)
  assumption_rows = tuple(assumption_rows)
  if not all(isinstance(row, VesselDecisionSummaryRow) for row in decision_rows):
    raise TypeError("decision_rows must contain VesselDecisionSummaryRow values")
  if tuple(row.vessel_id for row in decision_rows) != ("v1", "v2", "v3"):
    raise ValueError("decision_rows must contain exactly v1, v2, and v3")
  if not all(isinstance(row, AssumptionSourceRow) for row in assumption_rows):
    raise TypeError("assumption_rows must contain AssumptionSourceRow values")

  workbook = Workbook()
  decision_sheet = workbook.active
  decision_sheet.title = "Karar Özeti"
  assumptions_sheet = workbook.create_sheet("Varsayımlar")

  decision_sheet.append(DECISION_HEADERS)
  for row in decision_rows:
    decision_sheet.append((
        f"{row.vessel_id.upper()} — {row.vessel_name.split(' (', 1)[0]}",
        row.passenger_capacity,
        row.selected_cruise_speed_knots,
        row.battery_capacity_kwh,
        row.daily_propulsion_energy_kwh,
        row.solar_energy_contribution_kwh,
        row.net_grid_energy_requirement_kwh,
        (
            row.estimated_navigation_range_nm
            if row.estimated_navigation_range_nm is not None
            else "Mevcut değil"
        ),
        STATUS_LABELS[row.commission_compliance_status],
        row.investment_cost_tl,
        row.grant_amount_tl,
        row.net_investment_tl,
        row.annual_operating_saving_tl,
        (
            row.simple_payback_seasons
            if row.simple_payback_seasons is not None
            else "Mevcut değil"
        ),
        row.annual_co2_reduction_t,
    ))

  assumptions_sheet.append(ASSUMPTION_HEADERS)
  for row in assumption_rows:
    assumptions_sheet.append((
        row.parameter,
        row.current_value,
        row.source_type,
        row.description,
    ))

  _style_sheet(
      decision_sheet,
      {
          "A": 28, "B": 16, "C": 18, "D": 22, "E": 25,
          "F": 22, "G": 25, "H": 20, "I": 24, "J": 22,
          "K": 18, "L": 20, "M": 28, "N": 24, "O": 25,
      },
  )
  _style_sheet(
      assumptions_sheet,
      {"A": 34, "B": 28, "C": 34, "D": 68},
  )

  for column in ("C", "D", "E", "F", "G", "H", "N", "O"):
    for cell in decision_sheet[column][1:]:
      if isinstance(cell.value, (int, float)):
        cell.number_format = "#,##0.0"
  for column in ("J", "K", "L", "M"):
    for cell in decision_sheet[column][1:]:
      cell.number_format = '₺#,##0'

  output = BytesIO()
  workbook.save(output)
  return output.getvalue()
