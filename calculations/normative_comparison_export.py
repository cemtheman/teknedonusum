"""Excel and CSV exports for the immutable normative comparison contract."""

import csv
from io import BytesIO, StringIO

from openpyxl import Workbook

from calculations.decision_export import _style_sheet
from models.normative_vessel_comparison import NormativeVesselComparisonResult


COMPARISON_SHEET_NAME = "Normatif Karşılaştırma"
ASSUMPTIONS_SHEET_NAME = "Varsayımlar ve Kapsam"

COMPARISON_HEADERS = (
    "Tekne",
    "Tekne tipi",
    "Yolcu kapasitesi",
    "Hizmet hızı [kn]",
    "Motor gücü min [kW]",
    "Motor gücü referans [kW]",
    "Motor gücü max [kW]",
    "Günlük tahrik enerjisi min [kWh/gün]",
    "Günlük tahrik enerjisi referans [kWh/gün]",
    "Günlük tahrik enerjisi max [kWh/gün]",
    "Nominal batarya min [kWh]",
    "Nominal batarya referans [kWh]",
    "Nominal batarya max [kWh]",
    "Motor + batarya maliyeti min [EUR]",
    "Motor + batarya maliyeti referans [EUR]",
    "Motor + batarya maliyeti max [EUR]",
    "Metodoloji durumu",
    "Preliminary",
    "External validation",
)

SCOPE_LABELS = {
    "market_envelope_power_sizing": "Market-envelope based power sizing",
    "not_manufacturer_certified": "Non-certified",
    "not_sea_trial_validated": "Not sea-trial validated",
    "propulsion_energy_only": "Propulsion energy only",
    "auxiliary_and_hotel_loads_excluded": "Auxiliary/hotel loads excluded",
    "defined_motor_and_battery_cost_baseline_only": (
        "Cost scope: motor + battery"
    ),
    "solar_and_charging_infrastructure_excluded": (
        "Solar and charging infrastructure excluded"
    ),
}


def _validate(comparison):
  if not isinstance(comparison, NormativeVesselComparisonResult):
    raise TypeError("comparison must be a NormativeVesselComparisonResult")


def _comparison_values(row):
  return (
      row.vessel_id.upper(),
      row.vessel_type,
      row.passenger_capacity,
      row.selected_speed_knots,
      row.min_installed_mechanical_power_kw,
      row.reference_installed_mechanical_power_kw,
      row.max_installed_mechanical_power_kw,
      row.min_daily_propulsion_energy_kwh,
      row.reference_daily_propulsion_energy_kwh,
      row.max_daily_propulsion_energy_kwh,
      row.min_nominal_battery_capacity_kwh,
      row.reference_nominal_battery_capacity_kwh,
      row.max_nominal_battery_capacity_kwh,
      row.min_propulsion_system_cost,
      row.reference_propulsion_system_cost,
      row.max_propulsion_system_cost,
      row.methodology_status,
      row.preliminary_only,
      row.externally_validated,
  )


def _assumption_rows(comparison):
  assumptions = comparison.assumptions
  first = comparison.rows[0]
  rows = [
      ("Export type", "normative vessel comparison", "Export metadata"),
      ("Generated from", "NormativeVesselComparisonResult", "Data contract"),
      ("Selected speed [kn]", comparison.selected_speed_knots, "Common input"),
      ("Profile/method version", first.profile_version, "Normative profile"),
      ("Methodology status", first.methodology_status, "Preliminary status"),
      ("Validation status", first.validation_status, "Certification scope"),
      ("Motor efficiency", assumptions.motor_efficiency, "Common assumption"),
      (
          "Operating hours/day",
          assumptions.operating_hours_per_day,
          "Common assumption",
      ),
      ("Duty cycle", assumptions.duty_cycle, "Common assumption"),
      (
          "Effective powered hours/day",
          assumptions.effective_powered_hours_per_day,
          "Common assumption",
      ),
      (
          "Usable energy fraction",
          assumptions.usable_energy_fraction,
          "Common assumption",
      ),
      ("Reserve fraction", assumptions.reserve_fraction, "Common assumption"),
      (
          "Effective usable fraction",
          assumptions.effective_usable_energy_fraction,
          "Common assumption",
      ),
      ("Currency", comparison.currency, "Common output currency"),
  ]
  rows.extend(
      (
          f"{row.vessel_id.upper()} motor multiplier",
          row.motor_system_multiplier,
          "Vessel-specific existing cost assumption",
      )
      for row in comparison.rows
  )
  rows.extend(
      ("Model scope", SCOPE_LABELS.get(item, item), item)
      for item in first.limitation_ids
  )
  return tuple(rows)


def build_normative_comparison_xlsx(comparison) -> bytes:
  """Serialize contract values without recalculating sizing outputs."""
  _validate(comparison)
  workbook = Workbook()
  comparison_sheet = workbook.active
  comparison_sheet.title = COMPARISON_SHEET_NAME
  assumptions_sheet = workbook.create_sheet(ASSUMPTIONS_SHEET_NAME)

  comparison_sheet.append(COMPARISON_HEADERS)
  for row in comparison.rows:
    comparison_sheet.append(_comparison_values(row))

  assumptions_sheet.append(("Parametre", "Değer", "Açıklama"))
  for row in _assumption_rows(comparison):
    assumptions_sheet.append(row)

  _style_sheet(
      comparison_sheet,
      {chr(65 + index): width for index, width in enumerate(
          (12, 30, 18, 18, 22, 27, 22, 30, 36, 30, 25, 30, 25,
           32, 38, 32, 36, 16, 22)
      )},
  )
  _style_sheet(assumptions_sheet, {"A": 34, "B": 44, "C": 48})

  for column in range(4, 14):
    for cell in comparison_sheet.iter_cols(
        min_col=column,
        max_col=column,
        min_row=2,
    ):
      for item in cell:
        item.number_format = "#,##0.00"
  for column in range(14, 17):
    for cell in comparison_sheet.iter_cols(
        min_col=column,
        max_col=column,
        min_row=2,
    ):
      for item in cell:
        item.number_format = '#,##0.00 "EUR"'

  output = BytesIO()
  workbook.save(output)
  return output.getvalue()


def build_normative_comparison_csv(comparison) -> bytes:
  """Serialize only the main raw comparison table as UTF-8 CSV."""
  _validate(comparison)
  output = StringIO(newline="")
  writer = csv.writer(output)
  writer.writerow(COMPARISON_HEADERS)
  writer.writerows(_comparison_values(row) for row in comparison.rows)
  return output.getvalue().encode("utf-8-sig")
