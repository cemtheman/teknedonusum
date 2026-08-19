from io import BytesIO

import pytest
from openpyxl import load_workbook

from calculations.assumptions_transparency import (
    build_assumptions_transparency,
)
from calculations.decision_export import build_decision_summary_xlsx
from calculations.decision_summary import build_vessel_decision_summary
from calculations.economic_comparison import build_vessel_economic_comparison
from calculations.vessel_comparison import build_vessel_technical_comparison
from config.vessel_factory import build_vessel_specs
from models.inputs import SimulationInputs


def current_export():
  inputs = SimulationInputs(
      count_v1=50,
      count_v2=50,
      count_v3=40,
      count_v4_24=30,
      count_v4_32=20,
      cost_eur_v1=108100,
      cost_eur_v2=144140,
      cost_eur_v3=180180,
      eur_rate=55.5,
      diesel_price=81.81,
      elec_price=3.5,
      operating_days=180,
      sun_hours=8.0,
      daily_miles=35.0,
      cruise_speed=6.0,
  )
  specs = build_vessel_specs(
      inputs.cost_eur_v1,
      inputs.cost_eur_v2,
      inputs.cost_eur_v3,
      inputs.eur_rate,
  )
  technical = build_vessel_technical_comparison(
      specs,
      inputs.cruise_speed,
      inputs.daily_miles,
      inputs.sun_hours,
  )
  economic = build_vessel_economic_comparison(
      specs,
      inputs.cruise_speed,
      inputs.daily_miles,
      inputs.sun_hours,
      inputs.operating_days,
      inputs.elec_price,
      inputs.diesel_price,
      inputs.eur_rate,
  )
  decisions = build_vessel_decision_summary(technical, economic)
  assumptions = build_assumptions_transparency(inputs, specs, True, False)
  content = build_decision_summary_xlsx(decisions, assumptions)
  return load_workbook(BytesIO(content)), decisions, assumptions


def test_workbook_contains_exact_intended_sheets_and_vessel_rows():
  workbook, _, _ = current_export()

  assert workbook.sheetnames == ["Karar Özeti", "Varsayımlar"]
  sheet = workbook["Karar Özeti"]
  assert sheet.max_row == 4
  assert [sheet.cell(row, 1).value for row in range(2, 5)] == [
      "V1 — Tip 1: 12m Monohull",
      "V2 — Tip 2: 13.5m Katamaran",
      "V3 — Tip 3: 14m Katamaran",
  ]


def test_current_values_flow_into_export_as_numeric_cells():
  workbook, decisions, _ = current_export()
  sheet = workbook["Karar Özeti"]

  assert sheet["C2"].value == 6.0
  assert sheet["D2"].value == 80
  assert sheet["E2"].value == decisions[0].daily_energy_requirement_kwh
  assert sheet["J2"].value == 5999550
  assert sheet["M4"].value == pytest.approx(
      decisions[2].annual_operating_saving_tl
  )


def test_unavailable_values_remain_explicit():
  workbook, _, _ = current_export()
  sheet = workbook["Karar Özeti"]

  assert sheet["H3"].value == "Mevcut değil"
  assert sheet["H4"].value == "Mevcut değil"
  assert sheet["I3"].value == "Henüz değerlendirilmedi"
  assert sheet["I4"].value == "Henüz değerlendirilmedi"


def test_all_current_assumption_rows_are_exported():
  workbook, _, assumptions = current_export()
  sheet = workbook["Varsayımlar"]

  assert sheet.max_row == len(assumptions) + 1
  assert [cell.value for cell in sheet[1]] == [
      "Parametre",
      "Mevcut değer",
      "Kaynak türü",
      "Açıklama",
  ]
  exported_parameters = [
      sheet.cell(row, 1).value for row in range(2, sheet.max_row + 1)
  ]
  assert exported_parameters == [row.parameter for row in assumptions]
  assert sheet.cell(5, 3).value == "Canlı piyasa verisi — canlı"
  assert sheet.cell(6, 3).value == "Canlı piyasa verisi — statik yedek"


def test_workbook_has_simple_professional_formatting():
  workbook, _, _ = current_export()
  for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    assert all(cell.font.bold for cell in sheet[1])
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == sheet.dimensions
