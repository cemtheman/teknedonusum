from copy import deepcopy
from dataclasses import replace
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from calculations.assumptions_transparency import build_assumptions_transparency
from calculations.decision_export import build_decision_summary_xlsx
from calculations.decision_summary import build_vessel_decision_summary
from calculations.economic_comparison import build_vessel_economic_comparison
from calculations.fleet import calculate_fleet
from calculations.vessel_comparison import build_vessel_technical_comparison
from config.vessel_factory import build_vessel_specs
from models.inputs import SimulationInputs
from services import market_data


def vessel_specs():
  return build_vessel_specs(108100, 144140, 180180, 55.5)


def inputs(speed=4.0, miles=0.0, sun=0.0, days=30):
  return SimulationInputs(
      count_v1=0,
      count_v2=0,
      count_v3=0,
      count_v4_24=0,
      count_v4_32=0,
      cost_eur_v1=108100,
      cost_eur_v2=144140,
      cost_eur_v3=180180,
      eur_rate=55.5,
      diesel_price=81.81,
      elec_price=3.0,
      operating_days=days,
      sun_hours=sun,
      daily_miles=miles,
      cruise_speed=speed,
  )


@pytest.mark.parametrize(
    ("speed", "miles", "sun", "days"),
    [(4.0, 0.0, 0.0, 30), (10.0, 60.0, 12.0, 360)],
)
def test_operating_boundaries_build_comparisons_without_runtime_errors(
    speed,
    miles,
    sun,
    days,
):
  specs = vessel_specs()
  technical = build_vessel_technical_comparison(specs, speed, miles, sun)
  economic = build_vessel_economic_comparison(
      specs, speed, miles, sun, days, 3.0, 81.81, 55.5
  )
  decision = build_vessel_decision_summary(technical, economic)

  assert len(decision) == 3
  assert all(row.commission_compliance_status is None for row in decision)
  assert decision[1].estimated_navigation_range_nm is None
  assert decision[2].estimated_navigation_range_nm is None
  if miles == 0:
    assert all(row.daily_propulsion_energy_kwh == 0.0 for row in decision)
    assert all(row.net_grid_energy_requirement_kwh == 0.0 for row in decision)
  if sun == 0:
    assert all(row.solar_energy_contribution_kwh == 0.0 for row in decision)


def test_zero_vessel_counts_return_zero_fleet_totals_without_division_by_zero():
  specs = vessel_specs()
  specs["v4_24"] = deepcopy(specs["v1"])
  specs["v4_32"] = deepcopy(specs["v2"])
  counts = {key: 0 for key in specs}

  result = calculate_fleet(specs, counts, 4.0, 0.0, 0.0, 30)

  assert result.total_vessels == 0
  assert result.fleet_daily_brut_kwh == 0.0
  assert result.fleet_daily_solar_kwh == 0.0
  assert result.fleet_daily_grid_kwh == 0.0
  assert result.solar_coverage_ratio == 0.0


def test_xlsx_handles_zero_operation_and_all_unavailable_values():
  current_inputs = inputs()
  specs = vessel_specs()
  technical = build_vessel_technical_comparison(specs, 4.0, 0.0, 0.0)
  economic = build_vessel_economic_comparison(
      specs, 4.0, 0.0, 0.0, 30, 3.0, 81.81, 55.5
  )
  decision = tuple(
      replace(row, simple_payback_seasons=None)
      for row in build_vessel_decision_summary(technical, economic)
  )
  assumptions = build_assumptions_transparency(
      current_inputs, specs, False, False
  )

  content = build_decision_summary_xlsx(decision, assumptions)
  sheet = load_workbook(BytesIO(content))["Karar Özeti"]

  assert [sheet.cell(row, 5).value for row in range(2, 5)] == [0, 0, 0]
  assert [sheet.cell(row, 6).value for row in range(2, 5)] == [0, 0, 0]
  assert [sheet.cell(row, 7).value for row in range(2, 5)] == [0, 0, 0]
  assert [sheet.cell(row, 9).value for row in range(2, 5)] == [
      "Henüz değerlendirilmedi",
      "Henüz değerlendirilmedi",
      "Henüz değerlendirilmedi",
  ]
  assert [sheet.cell(row, 14).value for row in range(2, 5)] == [
      "Mevcut değil",
      "Mevcut değil",
      "Mevcut değil",
  ]


@pytest.mark.parametrize(
    ("fetcher_name", "expected_value"),
    [("fetch_tcmb_eur", 55.50), ("fetch_aytemiz_diesel", 81.81)],
)
def test_market_service_failure_returns_static_fallback(
    monkeypatch,
    fetcher_name,
    expected_value,
):
  def fail_request(*args, **kwargs):
    raise OSError("service unavailable")

  monkeypatch.setattr(market_data.urllib.request, "urlopen", fail_request)
  fetcher = getattr(market_data, fetcher_name)
  fetcher.clear()

  assert fetcher() == (expected_value, False)


def test_runtime_dependencies_cover_imported_third_party_packages():
  requirements = Path("requirements.txt").read_text(encoding="utf-8").splitlines()
  package_names = {line.split(">=", 1)[0] for line in requirements}

  assert {"streamlit", "pandas", "openpyxl"} <= package_names
