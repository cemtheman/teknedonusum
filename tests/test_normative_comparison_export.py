import csv
from io import BytesIO, StringIO

import pytest
from openpyxl import load_workbook

from calculations.normative_comparison_export import (
    ASSUMPTIONS_SHEET_NAME,
    COMPARISON_HEADERS,
    COMPARISON_SHEET_NAME,
    build_normative_comparison_csv,
    build_normative_comparison_xlsx,
)
from calculations.normative_vessel_comparison import (
    build_normative_vessel_comparison,
)


def exported(speed_knots):
  comparison = build_normative_vessel_comparison(speed_knots)
  content = build_normative_comparison_xlsx(comparison)
  return comparison, load_workbook(BytesIO(content))


@pytest.mark.parametrize("speed_knots", (6.0, 7.0, 8.0, 9.0, 10.0))
def test_workbook_propagates_comparison_rows_without_recalculation(speed_knots):
  comparison, workbook = exported(speed_knots)

  assert workbook.sheetnames == [COMPARISON_SHEET_NAME, ASSUMPTIONS_SHEET_NAME]
  sheet = workbook[COMPARISON_SHEET_NAME]
  assert sheet.max_row == 4
  assert tuple(cell.value for cell in sheet[1]) == COMPARISON_HEADERS
  assert tuple(sheet.cell(row, 1).value for row in range(2, 5)) == (
      "V1",
      "V2",
      "V3",
  )
  for index, source in enumerate(comparison.rows, start=2):
    actual = tuple(sheet.cell(index, column).value for column in range(3, 17))
    expected = (
        source.passenger_capacity,
        source.selected_speed_knots,
        source.min_installed_mechanical_power_kw,
        source.reference_installed_mechanical_power_kw,
        source.max_installed_mechanical_power_kw,
        source.min_daily_propulsion_energy_kwh,
        source.reference_daily_propulsion_energy_kwh,
        source.max_daily_propulsion_energy_kwh,
        source.min_nominal_battery_capacity_kwh,
        source.reference_nominal_battery_capacity_kwh,
        source.max_nominal_battery_capacity_kwh,
        source.min_propulsion_system_cost,
        source.reference_propulsion_system_cost,
        source.max_propulsion_system_cost,
    )
    assert actual == pytest.approx(expected)


def test_numeric_cells_and_formats_remain_numeric():
  _, workbook = exported(8.0)
  sheet = workbook[COMPARISON_SHEET_NAME]

  for row in range(2, 5):
    for column in range(3, 17):
      assert isinstance(sheet.cell(row, column).value, (int, float))
    assert sheet.cell(row, 6).number_format == "#,##0.00"
    assert sheet.cell(row, 15).number_format == '#,##0.00 "EUR"'


def test_assumptions_metadata_status_scope_and_multipliers():
  comparison, workbook = exported(8.0)
  sheet = workbook[ASSUMPTIONS_SHEET_NAME]
  values = {
      sheet.cell(row, 1).value: sheet.cell(row, 2).value
      for row in range(2, sheet.max_row + 1)
      if sheet.cell(row, 1).value != "Model scope"
  }
  scopes = [
      sheet.cell(row, 2).value
      for row in range(2, sheet.max_row + 1)
      if sheet.cell(row, 1).value == "Model scope"
  ]

  assert values["Generated from"] == "NormativeVesselComparisonResult"
  assert values["Selected speed [kn]"] == 8.0
  assert values["Motor efficiency"] == comparison.assumptions.motor_efficiency
  assert values["Effective usable fraction"] == pytest.approx(0.72)
  assert values["Currency"] == "EUR"
  assert values["V1 motor multiplier"] == 1.0
  assert values["V2 motor multiplier"] == 1.2
  assert values["V3 motor multiplier"] == 1.2
  assert "Non-certified" in scopes
  assert "Propulsion energy only" in scopes
  assert "Auxiliary/hotel loads excluded" in scopes
  assert "Solar and charging infrastructure excluded" in scopes
  assert "Cost scope: motor + battery" in scopes


def test_csv_contains_only_raw_main_table_in_vessel_order():
  comparison = build_normative_vessel_comparison(9.0)
  content = build_normative_comparison_csv(comparison)
  rows = list(csv.reader(StringIO(content.decode("utf-8-sig"))))

  assert tuple(rows[0]) == COMPARISON_HEADERS
  assert [row[0] for row in rows[1:]] == ["V1", "V2", "V3"]
  assert [float(row[3]) for row in rows[1:]] == [9.0, 9.0, 9.0]
  assert [float(row[5]) for row in rows[1:]] == [
      row.reference_installed_mechanical_power_kw for row in comparison.rows
  ]


@pytest.mark.parametrize(
    "builder",
    (build_normative_comparison_xlsx, build_normative_comparison_csv),
)
def test_exporters_require_comparison_contract(builder):
  with pytest.raises(TypeError, match="NormativeVesselComparisonResult"):
    builder(object())
